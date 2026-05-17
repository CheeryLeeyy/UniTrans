import os
import re
import argparse
from glob import glob
from typing import Dict, Optional

import yaml
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


EGO_LIST_DEFAULT = ["m3", "m4", "m7", "m13", "m17", "m31", "m36", "m39"]
METRICS = ["ap_30", "ap_50", "ap_70"]


def list_subdirs(root_dir: str):
    root_dir = os.path.abspath(root_dir)
    subdirs = []
    for name in sorted(os.listdir(root_dir)):
        p = os.path.join(root_dir, name)
        if os.path.isdir(p):
            subdirs.append(p)
    return subdirs


def find_yaml_for_ego(folder: str, ego: str) -> Optional[str]:
    cand = []
    for ext in ("*.yaml", "*.yml"):
        cand.extend(glob(os.path.join(folder, ext)))
    if not cand:
        return None

    pat = re.compile(rf"ego{re.escape(ego)}(?!\d)", re.IGNORECASE)

    matched = []
    for p in cand:
        base = os.path.basename(p)
        if pat.search(base):
            matched.append(p)

    if not matched:
        return None

    matched.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return matched[0]


def _extract_metrics_from_text(text: str) -> Dict[str, float]:
    out = {m: 0.0 for m in METRICS}
    for m in METRICS:
        r = re.search(rf"{re.escape(m)}\s*[:=]\s*([0-9]*\.?[0-9]+)", text)
        if r:
            out[m] = float(r.group(1))
    return out


def read_metrics_from_yaml(yaml_path: str) -> Dict[str, float]:
    with open(yaml_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()

    try:
        data = yaml.safe_load(text)
    except Exception:
        return _extract_metrics_from_text(text)

    if isinstance(data, dict):
        out = {}
        ok = 0
        for m in METRICS:
            v = data.get(m, None)
            if isinstance(v, (int, float)):
                out[m] = float(v)
                ok += 1
        if ok == len(METRICS):
            return out

    return _extract_metrics_from_text(text)


def mean_excluding_zeros(vals):
    nz = [v for v in vals if v != 0]
    if not nz:
        return 0.0
    return sum(nz) / len(nz)


def make_mean_formula(row_idx: int, col_start: int, col_end: int) -> str:
    start = f"{get_column_letter(col_start)}{row_idx}"
    end = f"{get_column_letter(col_end)}{row_idx}"
    rg = f"{start}:{end}"
    return f'=IF(COUNTIF({rg},">0")=0,0,SUMIF({rg},">0")/COUNTIF({rg},">0"))'


def write_excel(results, ego_list, out_xlsx):
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    headers = ["Metric", "Method"] + [e.upper() for e in ego_list] + ["Mean"]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    for method, per_ego in results.items():
        for metric in METRICS:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=method)

            col0 = 3
            for j, ego in enumerate(ego_list):
                v = per_ego.get(ego, {}).get(metric, 0.0)
                ws.cell(row=row, column=col0 + j, value=float(v))

            col_start = 3
            col_end = 3 + len(ego_list) - 1
            ws.cell(row=row, column=col_end + 1, value=make_mean_formula(row, col_start, col_end))

            row += 1

    for r in range(2, row):
        for c in range(3, 3 + len(ego_list) + 1):
            ws.cell(row=r, column=c).number_format = "0.0000"

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    for i in range(3, 3 + len(ego_list) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)
    wb.save(out_xlsx)


def write_csv(results, ego_list, out_csv: str):
    os.makedirs(os.path.dirname(os.path.abspath(out_csv)), exist_ok=True)

    headers = ["Metric", "Method"] + [e.upper() for e in ego_list] + ["Mean"]
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)

        for method, per_ego in results.items():
            for metric in METRICS:
                row = [metric, method]
                vals = []
                for ego in ego_list:
                    v = float(per_ego.get(ego, {}).get(metric, 0.0))
                    vals.append(v)
                    row.append(f"{v:.4f}")          # 改这里
                m = mean_excluding_zeros(vals)
                row.append(f"{m:.4f}")              # 改这里
                w.writerow(row)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--root_dir", type=str,
                        # required=True
                        default="opencood/logs/baseline",
                        )
    parser.add_argument("--out_xlsx", type=str, default="results.xlsx")
    parser.add_argument("--out_csv", type=str, default="results.csv")
    parser.add_argument("--ego_list", type=str, default=",".join(EGO_LIST_DEFAULT))
    args = parser.parse_args()

    root_dir = os.path.abspath(args.root_dir)
    ego_list = [x.strip() for x in args.ego_list.split(",") if x.strip()]

    results = {}
    skipped_all_missing = 0

    for sub in list_subdirs(root_dir):
        method = os.path.basename(sub)

        per_ego = {}
        nonzero_any = False

        for ego in ego_list:
            ypath = find_yaml_for_ego(sub, ego)
            if not ypath:
                per_ego[ego] = {m: 0.0 for m in METRICS}
                continue

            mvals = read_metrics_from_yaml(ypath)
            per_ego[ego] = mvals
            if any(mvals[m] > 0 for m in METRICS):
                nonzero_any = True

        if not nonzero_any:
            skipped_all_missing += 1
            continue

        results[method] = per_ego

    if not results:
        print("No valid experiment folders found (all skipped).")
        return

    xls_path = os.path.join(root_dir, args.out_xlsx)
    csv_path = os.path.join(root_dir, args.out_csv)
    write_excel(results, ego_list, xls_path)
    write_csv(results, ego_list, csv_path)

    print(f"Saved: {xls_path}")
    print(f"Saved: {csv_path}")
    print(f"Skipped folders (all egos missing): {skipped_all_missing}")
    print(f"Kept folders: {len(results)}")


if __name__ == "__main__":
    main()